# -*- coding: utf-8 -*-
import pandas as pd
import os
import argparse
import sys
import warnings
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# 设置输出编码为UTF-8
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

# 忽略xlrd和openpyxl的警告信息
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=FutureWarning, module='xlrd')


def split_excel_file(input_file, output_dir, rows_per_file, copy_headers=True):
    try:
        # 验证输入文件
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"输入文件不存在: {input_file}")
        
        if not input_file.lower().endswith(('.xlsx', '.xls')):
            raise ValueError(f"不支持的文件格式: {input_file}，仅支持.xlsx和.xls格式")
        
        # 验证参数
        if rows_per_file <= 0:
            raise ValueError(f"每个文件的行数必须大于0，当前值: {rows_per_file}")
        
        print(f"开始读取Excel文件: {input_file}")
        
        # 检测文件格式并选择合适的处理方式
        if input_file.lower().endswith('.xls'):
            print("检测到.xls格式文件，使用统一的嗅探式读取")
            print(f"文件路径: {input_file}")
            print(f"文件大小: {os.path.getsize(input_file)} 字节")
            
            # 使用统一的嗅探式读取，自动兼容HTML格式的.xls文件
            try:
                from utils import ExcelFileProcessor
                df = ExcelFileProcessor.read_excel_with_optimization(input_file)
                print(f"成功读取.xls文件，共{len(df)}行数据")
                
                # 将DataFrame转换为openpyxl工作簿以保持格式处理的一致性
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                wb = Workbook()
                ws = wb.active
                
                # 写入数据到工作表
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                
                # 对于HTML格式文件，df已经包含所有行（包括表头），不需要额外加1
                total_rows_with_header = len(df)
                
            except Exception as e:
                print(f"读取.xls文件失败: {e}")
                print("请检查文件格式是否正确")
                sys.exit(1)
        else:
            # 使用openpyxl读取Excel文件(.xlsx格式)
            print("检测到.xlsx格式文件，使用openpyxl引擎")
            wb = load_workbook(input_file, read_only=True)
            ws = wb.active
            
            # 获取总行数（包含表头）
            total_rows_with_header = ws.max_row
        
        if total_rows_with_header == 0:
            print("警告：Excel文件为空")
            return
        
        print(f"文件总行数（含表头）：{total_rows_with_header}")
        
        # 分割计算时自动排除源文件第一行表头（默认第一行为表头）
        # 数据行数始终为总行数减1（排除表头行）
        data_rows = max(0, total_rows_with_header - 1)
        
        print(f"数据行数（排除表头）：{data_rows}")
        if copy_headers:
            print("启用表头复制：每个分割文件将包含原始表头信息")
        else:
            print("关闭表头复制：所有分割文件均不包含表头信息")
        
        if data_rows == 0:
            print("警告：没有数据行需要拆分")
            # 创建一个空文件
            os.makedirs(output_dir, exist_ok=True)
            base_name = os.path.splitext(os.path.basename(input_file))[0]
            output_file = os.path.join(output_dir, f'{base_name}Split1.xlsx')
            new_wb = Workbook()
            new_ws = new_wb.active
            # 如果要求复制表头，复制表头
            if copy_headers and total_rows_with_header >= 1:
                for cell in ws[1]:
                    new_ws[cell.coordinate].value = cell.value
                    # 复制格式
                    new_ws[cell.coordinate].font = cell.font.copy()
                    new_ws[cell.coordinate].border = cell.border.copy()
                    new_ws[cell.coordinate].fill = cell.fill.copy()
                    new_ws[cell.coordinate].number_format = cell.number_format
                    new_ws[cell.coordinate].protection = cell.protection.copy()
                    new_ws[cell.coordinate].alignment = cell.alignment.copy()
            new_wb.save(output_file)
            file_rows = 1 if copy_headers and total_rows_with_header >= 1 else 0
            print(f'已创建文件：{output_file}（总行数：{file_rows}）')
            return
        
    except FileNotFoundError as e:
        print(f"错误: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"参数错误: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        sys.exit(1)

    # 计算分割文件数量（按数据行数切分）
    num_files = (data_rows + rows_per_file - 1) // rows_per_file
    print(f"准备拆分为{num_files}个文件，每个文件最多{rows_per_file}行")
    
    # 创建输出目录（如果不存在）
    os.makedirs(output_dir, exist_ok=True)

    # 复制列宽到后续新建工作簿
    template_wb = Workbook()
    template_ws = template_wb.active
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        template_ws.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width
    
    # 分割并保存文件
    for i in range(num_files):
        try:
            # 计算当前文件的数据行范围（基于数据行索引，从0开始）
            data_start_idx = i * rows_per_file
            data_end_idx = min((i + 1) * rows_per_file, data_rows)
            
            # 在源文件中的实际行号（数据从第2行开始，行号从1开始）
            source_start_row = 2 + data_start_idx
            source_end_row = 1 + data_end_idx
            
            print(f"正在处理第{i+1}/{num_files}个文件...")
            print(f"数据范围：第{source_start_row}行到第{source_end_row}行")
            
            # 创建新工作簿
            new_wb = Workbook()
            new_ws = new_wb.active
        
            current_write_row = 1
            
            # 复制表头（如果启用）
            if copy_headers:
                for cell in ws[1]:
                    tgt = new_ws.cell(row=current_write_row, column=cell.column, value=cell.value)
                    tgt.font = cell.font.copy()
                    tgt.border = cell.border.copy()
                    tgt.fill = cell.fill.copy()
                    tgt.number_format = cell.number_format
                    tgt.protection = cell.protection.copy()
                    tgt.alignment = cell.alignment.copy()
                current_write_row = 2  # 表头占用第1行，数据从第2行开始

            # 复制数据行（如果有数据）
            if data_end_idx > data_start_idx:
                for r_idx, row in enumerate(ws.iter_rows(min_row=source_start_row, max_row=source_end_row), start=0):
                    for cell in row:
                        new_cell = new_ws.cell(row=current_write_row + r_idx, column=cell.column, value=cell.value)
                        new_cell.font = cell.font.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.number_format = cell.number_format
                        new_cell.protection = cell.protection.copy()
                        new_cell.alignment = cell.alignment.copy()
            
            # 获取源文件名（不含扩展名）
            base_name = os.path.splitext(os.path.basename(input_file))[0]
            
            # 保存为新的Excel文件，使用源文件名+Split+序号格式
            output_file = os.path.join(output_dir, f'{base_name}Split{i+1}.xlsx')
            new_wb.save(output_file)

            # 计算实际行数
            actual_data_rows = data_end_idx - data_start_idx
            total_file_rows = actual_data_rows + (1 if copy_headers else 0)
            print(f'已创建文件：{output_file}（数据行数：{actual_data_rows}，总行数：{total_file_rows}）')
            
            # 计算并显示进度
            progress = ((i + 1) / num_files) * 100
            print(f"进度：{progress:.1f}% ({i+1}/{num_files})")
        
        except Exception as e:
            print(f"错误：处理第{i+1}个文件时失败: {e}")
            raise

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='拆分Excel文件（保留格式）')
    parser.add_argument('--input', required=True, help='输入Excel文件路径')
    parser.add_argument('--output', required=True, help='输出目录路径')
    parser.add_argument('--rows', type=int, default=1000, help='每个文件的行数（默认：1000）')
    parser.add_argument('--copy_headers', action='store_true', help='是否在每个拆分文件中复制表头')
    
    args = parser.parse_args()
    
    split_excel_file(args.input, args.output, args.rows, args.copy_headers)