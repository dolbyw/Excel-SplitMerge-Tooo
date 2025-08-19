# -*- coding: utf-8 -*-
import pandas as pd
import os
import argparse
import sys
import warnings
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# 设置输出编码为UTF-8
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

# 忽略xlrd和openpyxl的警告信息
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=FutureWarning, module='xlrd')


def split_excel_file(input_file, output_dir, rows_per_file, copy_headers=True):
    try:
        # 验证输入文件
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"输入文件不存在: {input_file}")
        
        if not input_file.lower().endswith(('.xlsx', '.xls')):
            raise ValueError(f"不支持的文件格式: {input_file}，仅支持.xlsx和.xls格式")
        
        # 验证参数
        if rows_per_file <= 0:
            raise ValueError(f"每个文件的行数必须大于0，当前值: {rows_per_file}")
        
        print(f"开始读取Excel文件: {input_file}")
        
        # 检测文件格式并选择合适的处理方式
        if input_file.lower().endswith('.xls'):
            print("检测到.xls格式文件，使用pandas+xlrd引擎读取")
            print(f"文件路径: {input_file}")
            print(f"文件大小: {os.path.getsize(input_file)} 字节")
            
            # 对于.xls文件，先用pandas读取数据，然后用openpyxl创建新文件
            try:
                print("尝试使用xlrd引擎读取.xls文件...")
                df = pd.read_excel(input_file, sheet_name=0, engine='xlrd')
                print(f"成功读取.xls文件，共{len(df)}行数据")
                
                # 将DataFrame转换为openpyxl工作簿以保持格式处理的一致性
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                wb = Workbook()
                ws = wb.active
                
                # 写入数据到工作表
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                
                total_rows_with_header = len(df) + 1  # 数据行数 + 表头
                
            except Exception as e:
                print(f"使用xlrd引擎失败: {e}")
                print("尝试使用默认方法读取.xls文件...")
                try:
                    df = pd.read_excel(input_file, sheet_name=0)
                    print(f"成功使用默认方法读取.xls文件，共{len(df)}行数据")
                    
                    from openpyxl import Workbook
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    
                    wb = Workbook()
                    ws = wb.active
                    
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                    
                    total_rows_with_header = len(df) + 1
                    
                except Exception as e2:
                    print(f"所有方法都失败: {e2}")
                    print("请确保已安装正确版本的xlrd库 (xlrd==1.2.0)")
                    print("可以尝试运行: pip install xlrd==1.2.0")
                    sys.exit(1)
        else:
            # 使用openpyxl读取Excel文件(.xlsx格式)
            print("检测到.xlsx格式文件，使用openpyxl引擎")
            wb = load_workbook(input_file, read_only=True)
            ws = wb.active
            
            # 获取总行数（包含表头）
            total_rows_with_header = ws.max_row
        
        if total_rows_with_header == 0:
            print("警告：Excel文件为空")
            return
        
        print(f"文件总行数（含表头）：{total_rows_with_header}")
        
        # 计算分割文件数量
        if copy_headers:
            # 如果复制表头，数据行数 = 总行数 - 1
            data_rows = max(0, total_rows_with_header - 1)
        else:
            # 如果不复制表头，所有行都是数据行
            data_rows = total_rows_with_header
        
        if data_rows == 0:
            print("警告：没有数据行需要拆分")
            # 创建一个空文件
            os.makedirs(output_dir, exist_ok=True)
            base_name = os.path.splitext(os.path.basename(input_file))[0]
            output_file = os.path.join(output_dir, f'{base_name}Split1.xlsx')
            new_wb = Workbook()
            new_ws = new_wb.active
            # 如果要求复制表头，复制表头
            if copy_headers and total_rows_with_header >= 1:
                for cell in ws[1]:
                    new_ws[cell.coordinate].value = cell.value
            new_wb.save(output_file)
            print(f'已创建文件：{output_file}（行数：{0 if not copy_headers else 1}）')
            return
        
    except FileNotFoundError as e:
        print(f"错误: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"参数错误: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        sys.exit(1)

    # 计算分割文件数量（按数据行数切分）
    num_files = (data_rows + rows_per_file - 1) // rows_per_file
    print(f"准备拆分为{num_files}个文件，每个文件最多{rows_per_file}行")
    
    # 创建输出目录（如果不存在）
    os.makedirs(output_dir, exist_ok=True)

    # 复制列宽到后续新建工作簿
    template_wb = Workbook()
    template_ws = template_wb.active
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        template_ws.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width
    
    # 分割并保存文件
    for i in range(num_files):
        try:
            # 本文件的数据起止行（在源表中的行号，均为数据行，不包括表头）
            data_start = 2 + i * rows_per_file  # 第2行是第一条数据
            data_end = min(1 + (i + 1) * rows_per_file, total_rows_with_header)  # 结束行包含端
            
            print(f"正在处理第{i+1}/{num_files}个文件...")
            
            # 创建新工作簿
            new_wb = Workbook()
            new_ws = new_wb.active
        
            # 复制表头（始终放在第1行）
            if copy_headers:
                for cell in ws[1]:
                    tgt = new_ws.cell(row=1, column=cell.column, value=cell.value)
                    tgt.font = cell.font.copy()
                    tgt.border = cell.border.copy()
                    tgt.fill = cell.fill.copy()
                    tgt.number_format = cell.number_format
                    tgt.protection = cell.protection.copy()
                    tgt.alignment = cell.alignment.copy()

            # 将数据行从第2行开始写入目标（如果复制了表头）
            write_start_row = 2 if copy_headers else 1
            for r_idx, row in enumerate(ws.iter_rows(min_row=data_start, max_row=data_end), start=0):
                for cell in row:
                    new_cell = new_ws.cell(row=write_start_row + r_idx, column=cell.column, value=cell.value)
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()
            
            # 获取源文件名（不含扩展名）
            base_name = os.path.splitext(os.path.basename(input_file))[0]
            
            # 保存为新的Excel文件，使用源文件名+Split+序号格式
            output_file = os.path.join(output_dir, f'{base_name}Split{i+1}.xlsx')
            new_wb.save(output_file)

            # 数据行数为 data_end - data_start + 1
            data_count = max(0, data_end - data_start + 1)
            print(f'已创建文件：{output_file}（行数：{data_count}）')
            
            # 计算并显示进度
            progress = ((i + 1) / num_files) * 100
            print(f"进度：{progress:.1f}% ({i+1}/{num_files})")
        
        except Exception as e:
            print(f"错误：处理第{i+1}个文件时失败: {e}")
            raise

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='拆分Excel文件（保留格式）')
    parser.add_argument('--input', required=True, help='输入Excel文件路径')
    parser.add_argument('--output', required=True, help='输出目录路径')
    parser.add_argument('--rows', type=int, default=1000, help='每个文件的行数（默认：1000）')
    parser.add_argument('--copy_headers', action='store_true', help='是否在每个拆分文件中复制表头')
    
    args = parser.parse_args()
    
    split_excel_file(args.input, args.output, args.rows, args.copy_headers)