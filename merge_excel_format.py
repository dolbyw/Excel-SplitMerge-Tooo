import pandas as pd
import os
import argparse
import glob
import sys
import warnings
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# 设置输出编码为UTF-8
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

# 忽略xlrd和openpyxl的警告信息
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=FutureWarning, module='xlrd')


def merge_excel_files(input_dir, output_file, remove_duplicate_headers=False):
    try:
        # 验证输入目录
        if not os.path.exists(input_dir):
            raise FileNotFoundError(f"输入目录不存在: {input_dir}")
        
        if not os.path.isdir(input_dir):
            raise ValueError(f"输入路径不是目录: {input_dir}")
        
        # 验证输出文件路径
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            print(f"创建输出目录: {output_dir}")
        
        print(f"开始扫描目录: {input_dir}")
        # 获取所有Excel文件
        excel_files = glob.glob(os.path.join(input_dir, "*.xlsx")) + glob.glob(os.path.join(input_dir, "*.xls"))
        
        if not excel_files:
            raise ValueError(f"在目录 {input_dir} 中未找到Excel文件(.xlsx/.xls)")
        
        print(f"找到{len(excel_files)}个Excel文件")
        for idx, file in enumerate(excel_files, 1):
            print(f"  文件{idx}: {os.path.basename(file)}")
        
    except FileNotFoundError as e:
        print(f"错误: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"参数错误: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"初始化失败: {e}")
        sys.exit(1)
    
    # 创建新工作簿作为模板
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    
    # 复制第一个文件的格式作为模板
    try:
        # 使用统一的嗅探式读取处理第一个文件
        from utils import ExcelFileProcessor
        first_df = ExcelFileProcessor.read_excel_with_optimization(excel_files[0])
        print(f"第一个文件详细信息: {os.path.basename(excel_files[0])}")
        print(f"  原始行数: {len(first_df)}行（包含表头）")
        print(f"  列数: {len(first_df.columns)}列")
        
        # 创建模板工作表并写入第一个文件的数据
        from openpyxl.utils.dataframe import dataframe_to_rows
        rows_written = 0
        for r in dataframe_to_rows(first_df, index=False, header=True):
            merged_ws.append(r)
            rows_written += 1
        print(f"  写入行数: {rows_written}行（列名表头1行 + DataFrame数据{len(first_df)}行）")
        print(f"  说明: DataFrame有{len(first_df)}行是因为pandas已将原文件第1行作为列名处理")
        
        # 设置默认列宽
        for col_num in range(1, len(first_df.columns) + 1):
            col_letter = get_column_letter(col_num)
            merged_ws.column_dimensions[col_letter].width = 15
        
        template_ws = merged_ws  # 使用合并工作表作为模板
        
    except Exception as e:
        print(f"处理第一个文件失败: {e}")
        sys.exit(1)
    
    # 当前行号（第一个文件已经写入）
    # dataframe_to_rows(first_df, header=True) 输出: 1行列名表头 + len(first_df)行数据 = len(first_df)+1行
    # 所以下一行的位置是: len(first_df) + 1 + 1 = len(first_df) + 2
    current_row = rows_written + 1  # 已写入行数 + 下一行位置
    print(f"下一个文件数据将从第{current_row}行开始写入")
    
    # 合并剩余文件的数据
    total_files = len(excel_files)
    print(f"准备合并{total_files}个Excel文件（保留格式）")
    print(f"第一个文件已处理: {excel_files[0]}")
    
    # 从第二个文件开始处理
    for i, file in enumerate(excel_files[1:], 1):
        try:
            print(f"正在读取文件 ({i+1}/{total_files}): {file}")
            
            # 使用统一的嗅探式读取
            try:
                df_current = ExcelFileProcessor.read_excel_with_optimization(file)
                if df_current.empty:
                    print(f"警告：文件 {os.path.basename(file)} 为空，跳过")
                    continue
                print(f"文件详细信息: {os.path.basename(file)}")
                print(f"  DataFrame行数: {len(df_current)}行（pandas已将原文件第1行作为列名）")
                print(f"  列数: {len(df_current.columns)}列")
                print(f"  说明: 原文件有{len(df_current)+1}行，第1行被pandas作为列名处理")
            except Exception as e:
                print(f"  读取文件失败: {e}，跳过此文件")
                continue
        
                # 根据表头去重设置处理数据
            if remove_duplicate_headers:
                # 开启表头去重：只添加数据行，不添加列名
                # 因为pandas已将原文件第1行作为列名，DataFrame中的所有行都是纯数据行
                # 所以我们应该保留DataFrame的所有行
                if len(df_current) > 0:
                    data_to_add = df_current.copy()  # 保留所有数据行
                    print(f"  启用表头去重：保留所有DataFrame数据行（pandas已处理表头）")
                    print(f"  处理结果: 添加{len(data_to_add)}行数据（原文件{len(df_current)+1}行 - 1行表头 = {len(df_current)}行数据）")
                else:
                    # DataFrame为空
                    data_to_add = pd.DataFrame()
                    print(f"  启用表头去重：DataFrame为空，无数据可添加")
                    print(f"  处理结果: 跳过整个文件（原文件只有表头行）")
            else:
                # 关闭表头去重：将列名作为数据行添加，然后添加所有DataFrame数据
                header_row = pd.DataFrame([df_current.columns.tolist()], columns=df_current.columns)
                data_to_add = pd.concat([header_row, df_current], ignore_index=True)
                print(f"  关闭表头去重：添加列名作为数据行，然后添加所有DataFrame数据")
                print(f"  处理结果: 添加{len(data_to_add)}行（1行列名 + {len(df_current)}行DataFrame数据）")
            
            # 将数据写入合并工作表
            rows_copied = 0
            if not data_to_add.empty:
                for _, row in data_to_add.iterrows():
                    for col_idx, value in enumerate(row, 1):
                        merged_ws.cell(row=current_row, column=col_idx, value=value)
                    current_row += 1
                    rows_copied += 1
            
            print(f"成功处理文件: {os.path.basename(file)}")
            print(f"  实际写入行数: {rows_copied}行")
            print(f"  当前总行数: {current_row - 1}行")
            
            # 显示进度
            progress = ((i + 1) / total_files) * 100
            print(f"合并进度：{progress:.1f}% ({i+1}/{total_files})")
            print(f"---")
            
        except Exception as e:
            print(f"错误：读取文件 {file} 失败: {e}")
            continue
    
    try:
        # 确保输出文件为.xlsx格式（即使输入包含.xls文件）
        if not output_file.lower().endswith('.xlsx'):
            if output_file.lower().endswith('.xls'):
                output_file = output_file[:-4] + '.xlsx'
                print(f"输出文件格式已自动转换为.xlsx格式")
            elif not output_file.lower().endswith(('.xlsx', '.xls')):
                output_file += '.xlsx'
                print(f"输出文件已自动添加.xlsx扩展名")
        
        print(f"开始保存到文件: {output_file}")
        # 保存合并后的文件（openpyxl自动保存为.xlsx格式）
        merged_wb.save(output_file)
        total_rows = current_row - 1
        print(f"="*60)
        print(f"合并完成！数据完整性统计:")
        print(f"  处理文件数: {len(excel_files)}个")
        print(f"  合并总行数: {total_rows}行（包含表头和数据）")
        print(f"  输出文件: {os.path.basename(output_file)}")
        print(f"  文件大小: {os.path.getsize(output_file)} 字节")
        print(f"="*60)
        
    except Exception as e:
        print(f"错误: 合并或保存文件失败: {e}")
        sys.exit(1)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='合并Excel文件（保留格式）')
    parser.add_argument('--input_dir', required=True, help='输入Excel文件所在目录')
    parser.add_argument('--output_file', required=True, help='输出文件路径')
    parser.add_argument('--remove_duplicate_headers', action='store_true', help='是否移除重复的表头')
    
    args = parser.parse_args()
    
    merge_excel_files(args.input_dir, args.output_file, args.remove_duplicate_headers)